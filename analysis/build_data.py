"""Build the dashboard's data artefacts from the source Excel + attack-rate table:
  1. dashboard/src/data/dengue.json  — per-district metrics, state totals, monthly series
  2. dashboard/public/tamilnadu_districts.geojson — simplified boundaries with a
     canonical `district` property for joining to the data.
"""
import json
import openpyxl, pandas as pd, geopandas as gpd
from _paths import XLSX, ATTACK_RATES_CSV, TN_GEOJSON, ROOT

YEARS = [2024, 2025, 2026]
PARTIAL = {2026: "Jan–Jun"}
DASH = ROOT / "dashboard"
DATA_OUT = DASH / "src" / "data" / "dengue.json"
GEO_OUT = DASH / "public" / "tamilnadu_districts.geojson"

norm = lambda s: str(s).lower().replace("the ", "").replace(" ", "").replace("-", "")
# geoBoundaries shapeName -> our canonical district name
GB_TO_OURS = {"Chengalputtu": "Chengalpattu", "Kancheepuram": "Kanchipuram",
              "Thoothukkudi": "Tuticorin", "Viluppuram": "Villupuram"}

# ---- canonical district list + per-year cases from the totals sheet ----
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["2024-26 yrs"]
districts = []
cases = {}
for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    n = str(r[0]).strip() if r[0] else ""
    if n and n != "State Total":
        districts.append(n)
        cases[n] = {2024: r[1] or 0, 2025: r[2] or 0, 2026: r[3] or 0}
canon = {norm(d): d for d in districts}

# ---- deaths + monthly series from the month-wise sheets ----
deaths = {d: {} for d in districts}
monthly = {y: {} for y in YEARS}  # monthly[year][district] = {cases:[12], deaths:[12]}
for y in YEARS:
    ws = wb[f"Dengue-Month wise- {y}"]
    for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        name = str(r[1]).strip() if r[1] else ""
        key = norm(name)
        if key not in canon:
            continue
        d = canon[key]
        mc = [r[2 + 2 * i] or 0 for i in range(12)]   # monthly cases
        md = [r[3 + 2 * i] or 0 for i in range(12)]   # monthly deaths
        deaths[d][y] = r[27] or 0                      # Total deaths column
        monthly[y][d] = {"cases": mc, "deaths": md}

# ---- population + attack rate from the processed table ----
ar = pd.read_csv(ATTACK_RATES_CSV).set_index("District")

# ---- assemble district records ----
district_records = []
state = {y: {"cases": 0, "deaths": 0, "population": 0} for y in YEARS}
for d in sorted(districts):
    rec = {"district": d, "metrics": {}}
    for y in YEARS:
        c = int(cases[d][y])
        dd = int(deaths[d].get(y, 0))
        pop = int(ar.at[d, f"Pop_{y}"])
        a = float(ar.at[d, f"AR_{y}"])
        cfr = round(dd / c * 100, 2) if c else 0.0
        rec["metrics"][str(y)] = {"cases": c, "deaths": dd, "population": pop,
                                   "attackRate": a, "cfr": cfr}
        state[y]["cases"] += c
        state[y]["deaths"] += dd
        state[y]["population"] += pop
    district_records.append(rec)

state_totals = {}
for y in YEARS:
    s = state[y]
    state_totals[str(y)] = {
        "cases": s["cases"], "deaths": s["deaths"], "population": s["population"],
        "attackRate": round(s["cases"] / s["population"] * 100000, 1),
        "cfr": round(s["deaths"] / s["cases"] * 100, 2) if s["cases"] else 0.0,
    }

out = {
    "meta": {
        "source": "Integrated Health Information Platform (IHIP), Tamil Nadu",
        "populationSource": "Government of Tamil Nadu (Census 2011, projected)",
        "years": YEARS,
        "partial": {str(k): v for k, v in PARTIAL.items()},
    },
    "districts": district_records,
    "stateTotals": state_totals,
    "monthly": {str(y): monthly[y] for y in YEARS},
}
DATA_OUT.parent.mkdir(parents=True, exist_ok=True)
DATA_OUT.write_text(json.dumps(out, indent=2), encoding="utf-8")
print("Wrote", DATA_OUT, "-", len(district_records), "districts")

# ---- simplified GeoJSON with canonical district names ----
gdf = gpd.read_file(TN_GEOJSON)[["shapeName", "geometry"]].copy()
gdf["district"] = gdf["shapeName"].map(lambda s: GB_TO_OURS.get(s, s))
# normalise any remaining names to our canonical spelling
gdf["district"] = gdf["district"].map(lambda s: canon.get(norm(s), s))
gdf["geometry"] = gdf["geometry"].simplify(0.004, preserve_topology=True)
gdf = gdf[["district", "geometry"]]
GEO_OUT.parent.mkdir(parents=True, exist_ok=True)
if GEO_OUT.exists():
    GEO_OUT.unlink()
gdf.to_file(GEO_OUT, driver="GeoJSON")
matched = gdf["district"].isin(districts).sum()
print(f"Wrote {GEO_OUT} - {len(gdf)} polygons, {matched}/38 joined; "
      f"{GEO_OUT.stat().st_size//1024} KB")
