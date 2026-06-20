"""Clip the India ADM2 boundaries to Tamil Nadu's 38 districts and render a
district-wise dengue case-count choropleth (2024). Downloads the open
geoBoundaries source file on first run if it is not already present."""
import urllib.request, ssl
import openpyxl, geopandas as gpd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from _paths import XLSX, INDIA_GEOJSON, INDIA_GEOJSON_URL, TN_GEOJSON, MAPS


def ensure_india_geojson():
    if INDIA_GEOJSON.exists():
        return
    print("Downloading geoBoundaries India ADM2 (~48 MB, one-time)...")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(INDIA_GEOJSON_URL, headers={"User-Agent": "Mozilla/5.0"})
    data = urllib.request.urlopen(req, timeout=180, context=ctx).read()
    INDIA_GEOJSON.write_bytes(data)


# 1. District case totals
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["2024-26 yrs"]
data = {}
for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    name = str(r[0]).strip() if r[0] else ""
    if name and name != "State Total":
        data[name] = {"y2024": r[1], "y2025": r[2], "y2026": r[3]}

# 2. Alias: our district name -> geoBoundaries shapeName
alias = {"Chengalpattu": "Chengalputtu", "Kanchipuram": "Kancheepuram",
         "Tuticorin": "Thoothukkudi", "Villupuram": "Viluppuram"}
norm = lambda s: s.lower().replace("the ", "").replace(" ", "").replace("-", "")
val2024 = {norm(alias.get(d, d)): v["y2024"] for d, v in data.items()}

# 3. Load India ADM2, keep only the 38 TN districts, save clipped file
ensure_india_geojson()
gdf = gpd.read_file(INDIA_GEOJSON)
gdf["key"] = gdf["shapeName"].map(norm)
tn = gdf[gdf["key"].isin(set(val2024))].copy()
tn["cases2024"] = tn["key"].map(val2024)
tn.to_file(TN_GEOJSON, driver="GeoJSON")
print(f"TN polygons: {len(tn)} | districts with data: {tn['cases2024'].notna().sum()}/38")

# 4. Choropleth
fig, ax = plt.subplots(figsize=(10, 11))
tn.plot(column="cases2024", cmap="YlOrRd", linewidth=0.4, edgecolor="grey",
        legend=True, ax=ax,
        legend_kwds={"label": "Dengue cases (2024)", "shrink": 0.6})
for _, row in tn.iterrows():
    c = row.geometry.representative_point()
    ax.annotate(row["shapeName"], (c.x, c.y), fontsize=5, ha="center")
ax.set_title("Tamil Nadu — Dengue Cases by District, 2024 (Source: IHIP)", fontsize=13)
ax.axis("off")
plt.tight_layout()
out = MAPS / "choropleth_2024_cases.png"
plt.savefig(out, dpi=200, bbox_inches="tight")
print("Saved", out)
