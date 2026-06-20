import openpyxl, pandas as pd
from _paths import XLSX, ATTACK_RATES_CSV

# ---- 1. Official TN population (Census 2011 base), in LAKHS, keyed to OUR data names
pop_lakhs = {
 'Chennai':46.47,'Kanchipuram':11.66,'Chengalpattu':25.56,'Thiruvallur':37.28,
 'Cuddalore':26.06,'Villupuram':20.93,'Kallakurichi':13.70,'Vellore':16.14,
 'Ranipet':12.10,'Tirupathur':11.12,'Tiruvannamalai':24.65,'Salem':34.82,
 'Namakkal':17.27,'Dharmapuri':15.07,'Krishnagiri':18.80,'Erode':22.52,
 'Coimbatore':34.58,'Tiruppur':24.79,'The Nilgiris':7.35,'Tiruchirappalli':27.22,
 'Karur':10.64,'Perambalur':5.65,'Ariyalur':7.55,'Thanjavur':24.06,
 'Thiruvarur':12.64,'Pudukkottai':16.18,'Madurai':30.38,'Theni':12.46,
 'Dindigul':21.60,'Ramanathapuram':13.53,'Virudhunagar':19.42,'Sivaganga':13.39,
 'Tirunelveli':16.65,'Tenkasi':14.08,'Tuticorin':17.50,'Kanniyakumari':18.70,
 # Nagapattinam 16.16 combined -> split using Census 2011 figures
 'Nagapattinam':6.98,'Mayiladuthurai':9.18,
}
pop2011 = {k: v*100000 for k,v in pop_lakhs.items()}   # lakhs -> persons

# ---- 2. Project to case years (TN decadal growth 2001-2011 = 15.61% -> 1.461%/yr geometric)
r = 1.1561**0.1 - 1
factor = {2024:(1+r)**13, 2025:(1+r)**14, 2026:(1+r)**15}

# ---- 3. Read case totals
wb=openpyxl.load_workbook(XLSX, data_only=True); ws=wb['2024-26 yrs']
cases={}
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    n=str(row[0]).strip() if row[0] else ''
    if n and n!='State Total':
        cases[n]={2024:row[1],2025:row[2],2026:row[3]}

# ---- 4. Build table: population (projected) + attack rate per 100k
rows=[]
for d in cases:
    rec={'District':d,'Pop_2011':round(pop2011[d])}
    for y in (2024,2025,2026):
        p=pop2011[d]*factor[y]
        rec[f'Pop_{y}']=round(p)
        rec[f'Cases_{y}']=cases[d][y]
        rec[f'AR_{y}']=round(cases[d][y]/p*100000,1)
    rows.append(rec)
df=pd.DataFrame(rows).sort_values('AR_2024',ascending=False)
df.to_csv(ATTACK_RATES_CSV, index=False)
print('Projection factors:', {y:round(f,4) for y,f in factor.items()}, ' (annual growth %.3f%%)'%(r*100))
print('State pop projected 2024 (millions):', round(sum(pop2011[d]*factor[2024] for d in cases)/1e6,2))
print()
print(df[['District','Pop_2024','Cases_2024','AR_2024','Cases_2025','AR_2025']].head(12).to_string(index=False))
print('\nSaved', ATTACK_RATES_CSV)
